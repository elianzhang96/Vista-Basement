import datetime
from gsapy import GSA
import openpyxl
from openpyxl.utils import get_column_letter
import json
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, Font, Alignment, Border, Side, PatternFill
from copy import copy
from scipy.spatial import ConvexHull
import numpy as np
import matplotlib.pyplot as plt
import math
import os
from mpl_toolkits.mplot3d import Axes3D


def extract_from_gsa(assembly_list, case_list, model: GSA):

    dictionary = dict()

    for assembly in assembly_list:
        time = datetime.datetime.now().replace(microsecond=0)
        print(f"   - Retrieving results for Assembly {assembly}. Start time: ", time)
        dictionary[assembly] = dict()
        
        for case in case_list:
            for permutation in define_permutations(model, case):
                results = model.get_assembly_forces(assembly, permutation)
                dictionary[assembly][permutation] = {'GSA Results': results}

                # dictionary[assembly][permutation] = dict()
                # dictionary[assembly][permutation]['GSA Results'] = []
                # dictionary[assembly][permutation]['GSA Results'] = model.get_assembly_forces(assembly, permutation)
    
    return dictionary

def write_list_to_excel(sheet, data, row_start, number_style):
    # write data to openpyxl sheet
    for i in range(len(data)):
        for j in range(len(data[i])):
            c = sheet.cell(row=i+row_start, column=j+1)
            c.value = data[i][j]
            if isinstance(c.value, (int, float)) and j != 0:
                c.style = number_style
    
    return sheet


def write_results_to_excel(sheetname, gsa_model_name, results_dict, outputFolder_excel, outputFileName_excel, number_style, start_time):

    # add to Excel            
    wb = openpyxl.load_workbook(outputFolder_excel + "\\" + outputFileName_excel)

    for assembly in results_dict.keys():

        # Create a line by line output to print in excel, in a nested list format
        results_table = [['Assembly', 'Case', 'Position','Fx [kN]', 'Fy [kN]', 'Fz [kN]', '|Fyz| [kN]', 'Mxx [kNm]', 'Myy [kNm]', 'Mzz [kNm]', '|Myz| [kNm]']]

        for case in results_dict[assembly].keys():
            for position in range(len(results_dict[assembly][case]['GSA Results'])):

                pos = results_dict[assembly][case]['GSA Results'][position][0]
                fx = results_dict[assembly][case]['GSA Results'][position][1]
                fy = results_dict[assembly][case]['GSA Results'][position][2]
                fz = results_dict[assembly][case]['GSA Results'][position][3]
                fyz = results_dict[assembly][case]['GSA Results'][position][4]
                mxx = results_dict[assembly][case]['GSA Results'][position][5]
                myy = results_dict[assembly][case]['GSA Results'][position][6]
                mzz = results_dict[assembly][case]['GSA Results'][position][7]
                myz = results_dict[assembly][case]['GSA Results'][position][8]

                results_table.append([assembly, case, pos, fx, fy, fz, fyz, mxx, myy, mzz, myz])

        sheet = wb.create_sheet(sheetname + str(assembly))

        sheet = write_list_to_excel(sheet, results_table, 3, number_style) # Export the list in excel
        c = sheet.cell(row=1, column=1) # Write the gsa file name at the top of this raw results table
        c.value = f"From GSA file: {gsa_model_name}. Time: {start_time}"
        c = sheet.cell(row=2, column=1) # Write the gsa file name at the top of this raw results table
        c.value = f"Assembly force and moment results."

        # Modify column width for improved readibility
        for column_cells in sheet.columns:
            column_letter = (get_column_letter(column_cells[0].column))
            if column_cells[0].value != 0:
                sheet.column_dimensions[column_letter].width = 10
        
        print(f"   - Results exported to Excel for Assembly {assembly}. Time: ", datetime.datetime.now().replace(microsecond=0))
    
    print("Saving to ", outputFileName_excel, ". Time: ", datetime.datetime.now().replace(microsecond=0))
    wb.save(outputFolder_excel + "\\" + outputFileName_excel)
    print("Results exported to ", outputFileName_excel, ". Time: ", datetime.datetime.now().replace(microsecond=0))
    
    return results_table

def write_criticalresults_to_excel(outputFileName_excel, sheetname, gsa_model_name, results_crit_dict, outputFolder_excel, number_style, start_time):

    # add to Excel            
    wb = openpyxl.Workbook()

    sheet = wb.create_sheet(sheetname)
    
    # Create a line by line output to print in excel, in a nested list format
    results_table = [['Assembly', 'Case', 'Sub-beam', 'Position','Fx [kN]', 'Fy [kN]', 'Fz [kN]', '|Fyz| [kN]', 'Mxx [kNm]', 'Myy [kNm]', 'Mzz [kNm]', '|Myz| [kNm]']]

    for assembly in results_crit_dict.keys():

        for sub_assembly in results_crit_dict[assembly].keys():

            for case in results_crit_dict[assembly][sub_assembly].keys():

                for position in range(len(results_crit_dict[assembly][sub_assembly][case])):

                    pos = results_crit_dict[assembly][sub_assembly][case][position][0]
                    fx = results_crit_dict[assembly][sub_assembly][case][position][1]
                    fy = results_crit_dict[assembly][sub_assembly][case][position][2]
                    fz = results_crit_dict[assembly][sub_assembly][case][position][3]
                    fyz = results_crit_dict[assembly][sub_assembly][case][position][4]
                    mxx = results_crit_dict[assembly][sub_assembly][case][position][5]
                    myy = results_crit_dict[assembly][sub_assembly][case][position][6]
                    mzz = results_crit_dict[assembly][sub_assembly][case][position][7]
                    myz = results_crit_dict[assembly][sub_assembly][case][position][8]

                    results_table.append([assembly, case, sub_assembly, pos, fx, fy, fz, fyz, mxx, myy, mzz, myz])

    sheet = write_list_to_excel(sheet, results_table, 3, number_style) # Export the list in excel
    c = sheet.cell(row=1, column=1) # Write the gsa file name at the top of this raw results table
    c.value = f"From GSA file: {gsa_model_name}. Time: {start_time}"
    c = sheet.cell(row=2, column=1) # Write the gsa file name at the top of this raw results table
    c.value = "Outlier Assembly force and moment results."

    # Modify column width for improved readibility
    for column_cells in sheet.columns:
        column_letter = (get_column_letter(column_cells[0].column))
        if column_cells[0].value != 0:
            sheet.column_dimensions[column_letter].width = 10

    wb.worksheets.insert(1, sheet)

    print("Saving to ", outputFileName_excel, ". Time: ", datetime.datetime.now().replace(microsecond=0))
    wb.save(outputFolder_excel + "\\" + outputFileName_excel)
    print("Outlier results exported to ", outputFileName_excel, ". Time: ", datetime.datetime.now().replace(microsecond=0))
    
    return results_table

def define_permutations(model:GSA, case):

    case_type, case_ref, _ = model.parse_case_string(case)
    num_perm = model.case_num_perm(case_type, case_ref)

    return (case_type + str(case_ref) + "p" + str(p+1) for p in range(num_perm)) if num_perm > 1 else (case,)

def split_results_subbeams(assembly_results_dict, sub_beams):

    for assembly in assembly_results_dict.keys():
        for case in assembly_results_dict[assembly]:
            no_positions = len(assembly_results_dict[assembly][case]['GSA Results']) # check how many points there are per assembly
            assembly_results_dict[assembly][case][sub_beams[0]] = [] # place where results for the first third of the beam are stored
            assembly_results_dict[assembly][case][sub_beams[1]] = [] # place where results for the mid third of the beam are stored
            assembly_results_dict[assembly][case][sub_beams[2]] = [] # place where results for the last third of the beam are stored

            for pos in range(0, math.ceil(no_positions/3)):
                assembly_results_dict[assembly][case][sub_beams[0]].append(assembly_results_dict[assembly][case]['GSA Results'][pos])
            
            for pos in range(math.floor(no_positions/3), math.ceil(no_positions*2/3)):
                assembly_results_dict[assembly][case][sub_beams[1]].append(assembly_results_dict[assembly][case]['GSA Results'][pos])
            
            for pos in range(math.floor(no_positions*2/3), math.ceil(no_positions)):
                assembly_results_dict[assembly][case][sub_beams[2]].append(assembly_results_dict[assembly][case]['GSA Results'][pos])

    return assembly_results_dict

def find_outliers(assembly_results_dict, sub_beams, plot_folder):
    
    # Create a list of points (sets of Fx, Myy and Mzz) and a list of node IDs to identify the outliers
    # Compute the convex hull to get ignore the sets of forces that are not governing
    # Create a dictionary with all critical forces

    sub_assembly_crit_results_dict = dict() # dictionary with all critical forces for design

    for assembly in assembly_results_dict.keys():
        sub_assembly_crit_results_dict[assembly] = dict()
        print(f"   - Finding outliers for Assembly {assembly}. Time: ", datetime.datetime.now().replace(microsecond=0))

        for n in sub_beams:

            sub_assembly_crit_results_dict[assembly][n] = dict()
            convex_hull_points = []
            convex_hull_ID = []

            for case in assembly_results_dict[assembly]:
                sub_assembly_crit_results_dict[assembly][n][case] = []
                
                for point in assembly_results_dict[assembly][case][n]:
                    convex_hull_points.append([point[6],point[1], point[7]])
                    convex_hull_ID.append(str(assembly)+"_"+n+"_"+case+"_"+str(point[0])) # Way to keep track of which point is being tested in the convex hull algorithm
            
            # Compute convex hull
            hull = ConvexHull(convex_hull_points)

            # Find outliers (points outside the convex hull) and their corresponding node IDs
            outliers = []
            outlier_node_ids = []
            for i, (point, node_id) in enumerate(zip(convex_hull_points, convex_hull_ID)):
                if i in hull.vertices:
                    outliers.append(point)
                    outlier_node_ids.append(node_id)
            
            # Add them to a critical result dictionary for design further down the script
            for id in outlier_node_ids:
                split_id = id.split('_')
                # Find via the ID the corresponding set of forces that is a vertex (critical value). Add it to the critical result dictionary.
                sub_assembly_crit_results_dict[int(split_id[0])][split_id[1]][split_id[2]].append(assembly_results_dict[int(split_id[0])][split_id[2]]['GSA Results'][int(float(split_id[3]))])

            print(f"        Outliers found for {n}. Time: ", datetime.datetime.now().replace(microsecond=0))

            # Plot the outliers for each sub-beam
            plot_convexhull(convex_hull_points, hull, outliers, assembly, n, plot_folder)
            print(f"        Plots exported for {n}. Time: ", datetime.datetime.now().replace(microsecond=0))

        
    return sub_assembly_crit_results_dict

def plot_convexhull(convex_hull_points, hull, outliers, assembly, n, plot_folder):

    # Convert convex_hull_points to a NumPy array
    convex_hull_points = np.array(convex_hull_points)

    # Calculate the maximum absolute values of x, y, and z coordinates
    max_abs_x = np.max(np.abs(convex_hull_points[:, 0]))
    max_abs_y = np.max(np.abs(convex_hull_points[:, 1]))
    max_abs_z = np.max(np.abs(convex_hull_points[:, 2]))
    max_abs = max(max_abs_x, max_abs_y, max_abs_z)
    
    # Plot the convex hull and outliers for this sub-assembly
    fig = plt.figure(figsize=(14, 11))
    ax = fig.add_subplot(111, projection='3d')

    # Plot convex hull
    for simplex in hull.simplices:
        ax.plot(convex_hull_points[simplex, 0], convex_hull_points[simplex, 1], convex_hull_points[simplex, 2], 'k-')

    # Create a list of points that are not outliers
    filtered_points = [point for point in convex_hull_points if not any(np.array_equal(point, outlier) for outlier in outliers)]

    # Plot the outliers
    ax.scatter(np.array(outliers)[:, 0], np.array(outliers)[:, 1], np.array(outliers)[:, 2], c='orange', label='Outliers')

    # Plot the filtered points
    filtered_points = np.array(filtered_points)  # Convert to NumPy array
    ax.scatter(filtered_points[:, 0], filtered_points[:, 1], filtered_points[:, 2], c='b', label='Points', alpha=0.5)

    # Set labels and title
    ax.set_xlabel('Myy')
    ax.set_ylabel('Fx')
    ax.set_zlabel('Mzz')
    plt.title(f'Convex Hull & Outliers for {assembly}, Sub-beam {n}')

     # Set limits of the axes symmetrically around zero based on the maximum absolute value
    ax.set_xlim([-max_abs_x, max_abs_x])
    ax.set_ylim([-max_abs_y, max_abs_y])
    ax.set_zlim([-max_abs_z, max_abs_z])

    # Add axes through the origin
    ax.plot([-1.5 * max_abs, 1.5 * max_abs], [0, 0], [0, 0], 'k-', lw=1)  # X-axis
    ax.plot([0, 0], [-1.5 * max_abs, 1.5 * max_abs], [0, 0], 'k-', lw=1)  # Y-axis
    ax.plot([0, 0], [0, 0], [-1.5 * max_abs, 1.5 * max_abs], 'k-', lw=1)  # Z-axis

    # Set viewing angle to emphasize x and y axes
    ax.view_init(elev=30, azim=100)
    # Add legend
    plt.legend()

    # Save plot
    plot_filename = f"{assembly}_sub_beam_{n}_plot.png"
    plot_filepath = os.path.join(plot_folder, plot_filename)
    plt.savefig(plot_filepath)
    plt.close()

